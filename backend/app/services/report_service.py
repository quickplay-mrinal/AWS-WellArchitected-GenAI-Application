from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from typing import Dict
from datetime import datetime
import io

def generate_pdf_report(scan: Dict) -> bytes:
    """Generate PDF report from scan results"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    story = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=30,
    )
    story.append(Paragraph("AWS Well-Architected Assessment Report", title_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Scan Info
    info_data = [
        ["Scan Name:", scan.get('scan_name', 'N/A')],
        ["Scan ID:", scan.get('scan_id', 'N/A')],
        ["Status:", scan.get('status', 'N/A')],
        ["Created:", scan.get('created_at', 'N/A')],
        ["Completed:", scan.get('completed_at', 'N/A')],
    ]
    
    info_table = Table(info_data, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.grey),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Executive Summary
    story.append(Paragraph("Executive Summary", styles['Heading2']))
    story.append(Spacer(1, 0.1*inch))
    
    regions = scan.get('regions_scanned', [])
    story.append(Paragraph(f"Scanned {len(regions)} AWS regions", styles['Normal']))
    story.append(Spacer(1, 0.2*inch))
    
    # AI Recommendations
    story.append(PageBreak())
    story.append(Paragraph("AI-Powered Recommendations", styles['Heading2']))
    story.append(Spacer(1, 0.1*inch))
    
    recommendations = scan.get('ai_recommendations', 'No recommendations available')
    for line in recommendations.split('\n'):
        if line.strip():
            story.append(Paragraph(line, styles['Normal']))
            story.append(Spacer(1, 0.05*inch))
    
    # Resource Summary
    story.append(PageBreak())
    story.append(Paragraph("Resource Summary by Region", styles['Heading2']))
    story.append(Spacer(1, 0.1*inch))
    
    results = scan.get('results', {})
    for region, data in results.items():
        story.append(Paragraph(f"<b>Region: {region}</b>", styles['Heading3']))
        
        resource_data = [
            ["Resource Type", "Count"],
            ["EC2 Instances", str(data.get('ec2', {}).get('count', 0))],
            ["RDS Databases", str(data.get('rds', {}).get('count', 0))],
            ["Lambda Functions", str(data.get('lambda', {}).get('count', 0))],
            ["S3 Buckets", str(data.get('s3', {}).get('count', 0))],
        ]
        
        resource_table = Table(resource_data, colWidths=[3*inch, 2*inch])
        resource_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(resource_table)
        story.append(Spacer(1, 0.2*inch))
    
    doc.build(story)
    return buffer.getvalue()

def generate_excel_report(scan: Dict) -> bytes:
    """Generate Excel report from scan results"""
    wb = Workbook()
    
    # Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Scan Info
    ws_summary['A1'] = "AWS Well-Architected Assessment Report"
    ws_summary['A1'].font = Font(size=16, bold=True)
    
    ws_summary['A3'] = "Scan Name:"
    ws_summary['B3'] = scan.get('scan_name', 'N/A')
    ws_summary['A4'] = "Scan ID:"
    ws_summary['B4'] = scan.get('scan_id', 'N/A')
    ws_summary['A5'] = "Status:"
    ws_summary['B5'] = scan.get('status', 'N/A')
    ws_summary['A6'] = "Created:"
    ws_summary['B6'] = scan.get('created_at', 'N/A')
    ws_summary['A7'] = "Completed:"
    ws_summary['B7'] = scan.get('completed_at', 'N/A')
    
    # Resources Sheet
    ws_resources = wb.create_sheet("Resources")
    headers = ["Region", "EC2 Instances", "RDS Databases", "Lambda Functions", "S3 Buckets", "VPCs", "Security Groups"]
    
    for col, header in enumerate(headers, 1):
        cell = ws_resources.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    results = scan.get('results', {})
    row = 2
    for region, data in results.items():
        ws_resources.cell(row=row, column=1, value=region)
        ws_resources.cell(row=row, column=2, value=data.get('ec2', {}).get('count', 0))
        ws_resources.cell(row=row, column=3, value=data.get('rds', {}).get('count', 0))
        ws_resources.cell(row=row, column=4, value=data.get('lambda', {}).get('count', 0))
        ws_resources.cell(row=row, column=5, value=data.get('s3', {}).get('count', 0))
        ws_resources.cell(row=row, column=6, value=data.get('vpc', {}).get('vpcs', 0))
        ws_resources.cell(row=row, column=7, value=data.get('vpc', {}).get('security_groups', 0))
        row += 1
    
    # Recommendations Sheet
    ws_recommendations = wb.create_sheet("AI Recommendations")
    ws_recommendations['A1'] = "AI-Powered Recommendations"
    ws_recommendations['A1'].font = Font(size=14, bold=True)
    
    recommendations = scan.get('ai_recommendations', 'No recommendations available')
    ws_recommendations['A3'] = recommendations
    ws_recommendations['A3'].alignment = Alignment(wrap_text=True, vertical="top")
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
